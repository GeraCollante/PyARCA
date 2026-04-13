#!/usr/bin/python
# This program is free software; you can redistribute it and/or modify
# it under the terms of the GNU Lesser General Public License as published by the
# Free Software Foundation; either version 3, or (at your option) any later
# version.
#
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of MERCHANTIBILITY
# or FITNESS FOR A PARTICULAR PURPOSE.  See the GNU Lesser General Public License
# for more details.

"Módulo para manejo de archivos CSV (planillas de cálculo)"

from builtins import range, zip

from past.builtins import basestring

__author__ = "Mariano Reingart (reingart@gmail.com)"
__copyright__ = "Copyright (C) 2010 Mariano Reingart"
__license__ = "LGPL-3.0-or-later"

import csv
import os


def leer(fn="entrada.csv", delimiter=";"):
    "Analiza un archivo CSV y devuelve un diccionario (aplanado)"
    ext = os.path.splitext(fn)[1].lower()
    items = []
    if ext == ".csv":
        csvfile = open(fn, "rb")
        # deducir dialecto y delimitador
        try:
            dialect = csv.Sniffer().sniff(csvfile.read(256), delimiters=[";", ","])
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = delimiter
        csvfile.seek(0)
        csv_reader = csv.reader(csvfile, dialect)
        for row in csv_reader:
            r = []
            for c in row:
                if isinstance(c, basestring):
                    c = c.strip()
                r.append(c)
            items.append(r)
    elif ext == ".xlsx":
        # extraigo los datos de la planilla Excel
        from openpyxl import load_workbook

        wb = load_workbook(filename=fn)
        ws1 = wb.get_active_sheet()
        for row in ws1.rows:
            fila = []
            for cell in row:
                fila.append(cell.value)
            items.append(fila)
    return items
    # TODO: return desaplanar(items)


def aplanar(regs):
    "Convierte una estructura python en planilla CSV (PyRece)"

    from formato_xml import MAP_ENC

    filas = []
    for reg in regs:
        fila = {}

        # recorrer campos obligatorios:
        for k in MAP_ENC:
            fila[k] = reg.get(k)

        fila["forma_pago"] = reg.get("forma_pago", "")
        fila["pdf"] = reg.get("pdf", "")

        # datos adicionales (escalares):
        for k, v in list(reg.items()):
            if k not in MAP_ENC and isinstance(k, (basestring, int)):
                fila[k] = v

        # por compatibilidad con pyrece:
        if reg.get("cbte_nro"):
            fila["cbt_numero"] = reg["cbte_nro"]

        for i, det in enumerate(reg["detalles"]):
            li = i + 1
            fila.update(
                {
                    f"codigo{li}": det.get("codigo", ""),
                    f"descripcion{li}": det.get("ds", ""),
                    f"umed{li}": det.get("umed"),
                    f"cantidad{li}": det.get("qty"),
                    f"precio{li}": det.get("precio"),
                    f"importe{li}": det.get("importe"),
                    f"iva_id{li}": det.get("iva_id"),
                    f"imp_iva{li}": det.get("imp_iva"),
                    f"bonif{li}": det.get("bonif"),
                    f"numero_despacho{li}": det.get("despacho"),
                    f"dato_a{li}": det.get("dato_a"),
                    f"dato_b{li}": det.get("dato_b"),
                    f"dato_c{li}": det.get("dato_c"),
                    f"dato_d{li}": det.get("dato_d"),
                    f"dato_e{li}": det.get("dato_e"),
                }
            )
        for i, iva in enumerate(reg["ivas"]):
            li = i + 1
            fila.update(
                {
                    f"iva_id_{li}": iva["iva_id"],
                    f"iva_base_imp_{li}": iva["base_imp"],
                    f"iva_importe_{li}": iva["importe"],
                }
            )
        for i, tributo in enumerate(reg["tributos"]):
            li = i + 1
            fila.update(
                {
                    f"tributo_id_{li}": tributo["tributo_id"],
                    f"tributo_base_imp_{li}": tributo["base_imp"],
                    f"tributo_desc_{li}": tributo["desc"],
                    f"tributo_alic_{li}": tributo["alic"],
                    f"tributo_importe_{li}": tributo["importe"],
                }
            )
        for i, opcional in enumerate(reg.get("opcionales", [])):
            li = i + 1
            fila.update(
                {
                    f"opcional_id_{li}": opcional["opcional_id"],
                    f"opcional_valor_{li}": opcional["valor"],
                }
            )
        for i, cbte_asoc in enumerate(reg.get("cbtes_asoc", [])):
            li = i + 1
            fila.update(
                {
                    f"cbte_asoc_tipo_{li}": cbte_asoc["cbte_tipo"],
                    f"cbte_asoc_pto_vta_{li}": cbte_asoc["cbte_punto_vta"],
                    f"cbte_asoc_nro_{li}": cbte_asoc["cbte_nro"],
                    f"cbte_asoc_cuit_{li}": cbte_asoc["cbte_cuit"],
                    f"cbte_asoc_fecha_{li}": cbte_asoc["cbte_fecha"],
                }
            )

        filas.append(fila)

    cols = [
        "id",
        "tipo_cbte",
        "punto_vta",
        "cbt_numero",
        "fecha_cbte",
        "tipo_doc",
        "nro_doc",
        "moneda_id",
        "moneda_ctz",
        "imp_neto",
        "imp_iva",
        "imp_trib",
        "imp_op_ex",
        "imp_tot_conc",
        "imp_total",
        "concepto",
        "fecha_venc_pago",
        "fecha_serv_desde",
        "fecha_serv_hasta",
        "cae",
        "fecha_vto",
        "resultado",
        "motivo",
        "reproceso",
        "nombre",
        "domicilio",
        "localidad",
        "telefono",
        "categoria",
        "email",
        "numero_cliente",
        "numero_orden_compra",
        "condicion_frente_iva",
        "numero_cotizacion",
        "numero_remito",
        "obs_generales",
        "obs_comerciales",
    ]

    # filtro y ordeno las columnas
    keys_list = [k for f in filas for k in list(f.keys())]
    s = set(keys_list) - set(cols)
    cols = cols + list(s)

    ret = [cols]
    for fila in filas:
        ret.append([fila.get(k) for k in cols])

    return ret


def desaplanar(filas):
    "Dado una planilla, conviertir en estructura python"

    from formato_xml import MAP_ENC

    def max_li(colname):
        nums = [int(k[len(colname) :]) + 1 for k in filas[0] if k.startswith(colname)]
        if nums:
            tmp = max(nums)
        if nums and tmp:
            ##print "max_li(%s)=%s" % (colname, tmp)
            return tmp
        else:
            return 0

    regs = []
    for fila in filas[1:]:
        dic = dict([(filas[0][i], v) for i, v in enumerate(fila)])
        reg = {}

        # por compatibilidad con pyrece:
        reg["cbte_nro"] = dic["cbt_numero"]

        for k in MAP_ENC:
            if k in dic:
                reg[k] = dic.pop(k)

        reg["detalles"] = [
            {
                "codigo": (f"codigo{li}") in dic and dic.pop(f"codigo{li}") or None,
                "ds": (f"descripcion{li}") in dic
                and dic.pop(f"descripcion{li}")
                or None,
                "umed": (f"umed{li}") in dic and dic.pop(f"umed{li}") or None,
                "qty": (f"cantidad{li}") in dic
                and dic.pop(f"cantidad{li}")
                or None,
                "precio": (f"precio{li}") in dic and dic.pop(f"precio{li}") or None,
                "importe": (f"importe{li}") in dic
                and dic.pop(f"importe{li}")
                or None,
                "iva_id": (f"iva_id{li}") in dic and dic.pop(f"iva_id{li}") or None,
                "imp_iva": (f"imp_iva{li}") in dic
                and dic.pop(f"imp_iva{li}")
                or None,
                "bonif": (f"bonif{li}") in dic and dic.pop(f"bonif{li}") or None,
                "despacho": (f"numero_despacho{li}") in dic
                and dic.pop(f"numero_despacho{li}"),
                "dato_a": (f"dato_a{li}") in dic and dic.pop(f"dato_a{li}"),
                "dato_b": (f"dato_b{li}") in dic and dic.pop(f"dato_b{li}"),
                "dato_c": (f"dato_c{li}") in dic and dic.pop(f"dato_c{li}"),
                "dato_d": (f"dato_d{li}") in dic and dic.pop(f"dato_d{li}"),
                "dato_e": (f"dato_e{li}") in dic and dic.pop(f"dato_e{li}"),
            }
            for li in range(1, max_li("cantidad"))
            if dic[f"cantidad{li}"] is not None
        ]

        # descartar filas espurias vacias al final
        for det in reg["detalles"][::-1]:
            if any(det.values()):  # algun campo tiene dato termina
                break
            del reg["detalles"][-1]  # sino, borro último elemento

        reg["tributos"] = [
            {
                "tributo_id": dic.pop(f"tributo_id_{li}"),
                "desc": dic.pop(f"tributo_desc_{li}"),
                "base_imp": dic.pop(f"tributo_base_imp_{li}"),
                "alic": dic.pop(f"tributo_alic_{li}"),
                "importe": dic.pop(f"tributo_importe_{li}"),
            }
            for li in range(1, max_li("tributo_id_"))
            if dic[f"tributo_id_{li}"]
        ]

        reg["ivas"] = [
            {
                "iva_id": dic.pop(f"iva_id_{li}"),
                "base_imp": dic.pop(f"iva_base_imp_{li}"),
                "importe": dic.pop(f"iva_importe_{li}"),
            }
            for li in range(1, max_li("iva_id_"))
            if dic[f"iva_id_{li}"]
        ]

        reg["permisos"] = [
            {
                "id_permiso": dic.pop(f"id_permiso_{li}"),
                "dst_merc": dic.pop(f"dst_merc_{li}"),
            }
            for li in range(1, max_li("id_permiso_"))
            if dic[f"id_permiso_{li}"]
        ]

        reg["opcionales"] = [
            {
                "opcional_id": dic.pop(f"opcional_id_{li}"),
                "valor": dic.pop(f"opcional_valor_{li}"),
            }
            for li in range(1, max_li("opcional_id_"))
            if dic[f"opcional_id_{li}"]
        ]

        reg["cbtes_asoc"] = [
            {
                "cbte_tipo": dic.pop(f"cbte_asoc_tipo_{li}"),
                "cbte_punto_vta": dic.pop(f"cbte_asoc_pto_vta_{li}"),
                "cbte_nro": dic.pop(f"cbte_asoc_nro_{li}"),
                "cbte_cuit": dic.pop(f"cbte_asoc_cuit_{li}"),
                "cbte_fecha": dic.pop(f"cbte_asoc_fecha_{li}"),
            }
            for li in range(1, max_li("cbte_asoc_tipo_"))
            if dic[f"cbte_asoc_tipo_{li}"]
        ]

        reg["forma_pago"] = dic.pop("forma_pago")

        # agrego campos adicionales:
        reg["datos"] = [
            {
                "campo": campo,
                "valor": valor,
                "pagina": "",
            }
            for campo, valor in list(dic.items())
        ]

        regs.append(reg)

    return regs


def escribir(filas, fn="salida.csv", delimiter=";"):
    "Dado una lista de comprobantes (diccionarios), aplana y escribe"
    ext = os.path.splitext(fn)[1].lower()
    if ext == ".csv":
        f = open(fn, "wb")
        csv_writer = csv.writer(f, dialect="excel", delimiter=";")
        # TODO: filas = aplanar(regs)
        for fila in filas:
            # convertir a ISO-8859-1 (evita error de encoding de csv writer):
            fila = [
                celda.encode("latin1") if isinstance(celda, str) else celda
                for celda in fila
            ]
            csv_writer.writerow(fila)
        f.close()
    elif ext == ".xlsx":
        from openpyxl import Workbook

        wb = Workbook()
        ws1 = wb.get_active_sheet()
        for fila in filas:
            ws1.append(fila)
        wb.save(filename=fn)


# pruebas básicas
if __name__ == "__main__":
    ##import pdb; pdb.set_trace()
    filas = leer("facturas-wsfev1-bis.csv")
    regs1 = desaplanar(filas)
    print(filas)
    filas1 = aplanar(regs1)
    print(filas1)
    print(filas1 == filas)
    escribir(filas1, "facturas-wsfev1-bis-sal.csv")
    escribir(filas1, "facturas-wsfev1-bis-sal.xlsx")
    filas2 = leer("facturas-wsfev1-bis-sal.xlsx")
    for fila1, fila2 in zip(filas1, filas2):
        for celda1, celda2 in zip(fila1, fila2):
            if celda1 != celda2:
                print(celda1, celda2)
